import serial
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime

now = datetime.now()
current_time = now.strftime("%H:%M:%S")
ser = serial.Serial('COM4', 9600, timeout=1)
time.sleep(2)
wb = load_workbook('voltage.xlsx')
ws = wb.active
ws['A1'] = 'Time'
ws['B1'] = 'Voltage'

a=0
try:
    while a<30:
        if ser.in_waiting > 0:
            line = ser.readline().decode('utf-8').rstrip()
            ws.append([current_time,line])
            wb.save('voltage.xlsx')
            a+=1 
except KeyboardInterrupt:
    print("Exiting...")
finally:
    ser.close()
